"""
export_excel.py — exports the DB into a dated multi-sheet Excel file.

Sheet layout (all labels in Greek):
  1. Οχήματα ανά Ημέρα     — Col A (blue)=vehicle number sorted, Col B (green)=route
  2. Πλέγμα Δρομολογίων    — rows=04:00-23:55 in 5min steps, cols=routes,
                              cells=vehicle number that departed at that time
  3. Δρομολόγια & Στάσεις  — every trip with stop-level pass times
  4. Ημερήσια Στατιστικά   — daily route stats (actual vs scheduled, %, deviation)
  5. Αναχωρήσεις Οχημάτων  — vehicle departure events
  6. Αλλαγές Βάρδιας       — slot handoffs (shift changes)
  7. Γραμμές & Διαδρομές   — master reference

Usage:
    python scripts/export_excel.py [output.xlsx] [--date YYYY-MM-DD] [--days N] [--all]
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import db

# Colours
BLUE_FILL  = PatternFill("solid", start_color="4472C4")  # vehicle column
GREEN_FILL = PatternFill("solid", start_color="70AD47")  # route column
HEADER_FILL = PatternFill("solid", start_color="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
BODY_FONT   = Font(name="Calibri", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")


# ── Greek labels ──────────────────────────────────────────────────────────────

SHEET_NAMES = {
    "vehicles":    "Οχήματα ανά Ημέρα",
    "grid":        "Πλέγμα Δρομολογίων",
    "trips":       "Δρομολόγια & Στάσεις",
    "stats":       "Ημερήσια Στατιστικά",
    "departures":  "Αναχωρήσεις Οχημάτων",
    "handoffs":    "Αλλαγές Βάρδιας",
    "reference":   "Γραμμές & Διαδρομές",
}

COL_VEHICLE_NO  = "Αριθμός Οχήματος"
COL_ROUTE       = "Διαδρομή"
COL_LINE        = "Γραμμή"
COL_DATE        = "Ημερομηνία"
COL_DIRECTION   = "Κατεύθυνση"
COL_SLOT        = "Αρ. Σειράς"
COL_DEPARTED    = "Αναχώρηση (UTC)"
COL_ARRIVED     = "Άφιξη Τέρματος (UTC)"
COL_STOPS       = "Στάσεις"
COL_ACTUAL      = "Πραγματικά Δρομολόγια"
COL_SCHED       = "Προγραμματισμένα"
COL_PCT         = "Εκτέλεση %"
COL_DEVIATION   = "Μέση Καθυστέρηση (λεπτά)"
COL_VEHICLES_N  = "Διαφορετικά Οχήματα"
COL_SLOTS_N     = "Αρ. Σειρών"
COL_SCHED_DEP   = "Προγρ. Αναχώρηση"
COL_DEV_MINS    = "Καθυστέρηση (λεπτά)"
COL_OUTGOING    = "Εξερχόμενο Όχημα"
COL_INCOMING    = "Εισερχόμενο Όχημα"
COL_HANDOFF_T   = "Ώρα Αλλαγής (UTC)"
COL_GAP         = "Διάλειμμα (λεπτά)"
COL_STOP_ORDER  = "Σειρά Στάσης"
COL_STOP_CODE   = "Κωδικός Στάσης"
COL_STOP_NAME   = "Όνομα Στάσης"
COL_PASSED      = "Πέρασμα (UTC)"
COL_DIST        = "Απόσταση (μ)"
COL_METHOD      = "Μέθοδος"
COL_TRIP_ID     = "ID Δρομολογίου"
COL_DISTANCE_M  = "Απόσταση Διαδρομής (μ)"
COL_OUTBOUND    = "Εξερχόμενη"
COL_INBOUND     = "Εισερχόμενη"


def direction_gr(route_type: str | None) -> str:
    if route_type == "1":
        return COL_OUTBOUND
    if route_type == "2":
        return COL_INBOUND
    return route_type or ""


# ── Data loaders ──────────────────────────────────────────────────────────────

def load_vehicles_per_day(conn, service_date: str) -> pd.DataFrame:
    rows = conn.execute("""
        SELECT DISTINCT va.vehicle_no, r.line_code, r.descr AS route_descr,
               va.route_code
        FROM vehicle_activity va
        LEFT JOIN routes r ON r.route_code = va.route_code
        WHERE va.service_date = ?
        ORDER BY CAST(va.vehicle_no AS INTEGER), va.vehicle_no, va.route_code
    """, (service_date,)).fetchall()
    return pd.DataFrame([{
        COL_VEHICLE_NO: r["vehicle_no"],
        COL_ROUTE:      f"{r['line_code']} – {r['route_descr']}" if r["line_code"] else r["route_code"],
    } for r in rows])


def load_departure_grid(conn, service_date: str) -> pd.DataFrame:
    """
    Returns a DataFrame:
      index = time strings '04:00' … '23:55' (5-min steps)
      columns = route descriptions
      values = vehicle_no that departed at that time, or ''
    """
    # Time index
    times = []
    h, m = 4, 0
    while h < 24:
        times.append(f"{h:02d}:{m:02d}")
        m += 5
        if m == 60:
            m = 0
            h += 1

    # Get all departures for the day with their matched scheduled time
    rows = conn.execute("""
        SELECT t.vehicle_no,
               t.route_code,
               r.line_code,
               r.descr AS route_descr,
               sa.scheduled_departure
        FROM trips t
        JOIN slot_assignments sa ON sa.trip_id = t.id
        LEFT JOIN routes r ON r.route_code = t.route_code
        WHERE t.service_date = ?
        ORDER BY t.route_code, sa.scheduled_departure
    """, (service_date,)).fetchall()

    route_label_map: dict[str, str] = {}
    cell_map: dict[tuple[str, str], str] = {}  # (time_str, route_label) → vehicle_no

    for r in rows:
        route_label = (f"{r['line_code']} {r['route_descr']}"
                       if r["line_code"] else (r["route_descr"] or r["route_code"]))
        route_label_map[r["route_code"]] = route_label

        sched = r["scheduled_departure"]  # 'HH:MM:SS' or None
        if sched:
            time_key = sched[:5]  # 'HH:MM'
            if time_key in times:
                cell_map[(time_key, route_label)] = r["vehicle_no"]

    all_routes = sorted(set(route_label_map.values()))
    df = pd.DataFrame(index=times, columns=all_routes, dtype=str)
    df = df.fillna("")

    for (time_key, route_label), veh in cell_map.items():
        if route_label in df.columns and time_key in df.index:
            df.at[time_key, route_label] = veh

    df.index.name = "Ώρα"
    return df


def load_trips_with_stops(conn, service_date: str, days: int) -> pd.DataFrame:
    cutoff = (date.fromisoformat(service_date) - timedelta(days=days)).isoformat()
    rows = conn.execute("""
        SELECT t.id, t.service_date, t.vehicle_no,
               r.line_code, r.descr AS route_descr, r.route_type,
               t.started_at, t.ended_at, t.terminus_arrived_at,
               tst.stop_order, tst.stop_code,
               s.descr AS stop_name,
               tst.passed_at, tst.distance_m, tst.method
        FROM trips t
        LEFT JOIN routes r ON r.route_code = t.route_code
        LEFT JOIN trip_stop_times tst ON tst.trip_id = t.id
        LEFT JOIN stops s ON s.route_code = t.route_code
                          AND s.stop_order = tst.stop_order
        WHERE t.service_date >= ?
        ORDER BY t.service_date DESC, t.vehicle_no, t.started_at, tst.stop_order
    """, (cutoff,)).fetchall()

    return pd.DataFrame([{
        COL_TRIP_ID:    r["id"],
        COL_DATE:       r["service_date"],
        COL_VEHICLE_NO: r["vehicle_no"],
        COL_LINE:       r["line_code"],
        COL_ROUTE:      r["route_descr"],
        COL_DIRECTION:  direction_gr(r["route_type"]),
        COL_DEPARTED:   r["started_at"],
        COL_ARRIVED:    r["terminus_arrived_at"] or r["ended_at"],
        COL_STOP_ORDER: r["stop_order"],
        COL_STOP_CODE:  r["stop_code"],
        COL_STOP_NAME:  r["stop_name"],
        COL_PASSED:     r["passed_at"],
        COL_DIST:       r["distance_m"],
        COL_METHOD:     r["method"],
    } for r in rows])


def load_daily_stats(conn) -> pd.DataFrame:
    rows = conn.execute("""
        SELECT drs.service_date, r.line_code, r.descr AS route_descr,
               r.route_type, drs.actual_trip_count, drs.scheduled_trip_count,
               drs.completion_pct, drs.avg_deviation_mins,
               drs.distinct_vehicles, drs.slot_count
        FROM daily_route_stats drs
        LEFT JOIN routes r ON r.route_code = drs.route_code
        ORDER BY drs.service_date DESC, r.line_code
    """).fetchall()
    return pd.DataFrame([{
        COL_DATE:       r["service_date"],
        COL_LINE:       r["line_code"],
        COL_ROUTE:      r["route_descr"],
        COL_DIRECTION:  direction_gr(r["route_type"]),
        COL_ACTUAL:     r["actual_trip_count"],
        COL_SCHED:      r["scheduled_trip_count"],
        COL_PCT:        r["completion_pct"],
        COL_DEVIATION:  r["avg_deviation_mins"],
        COL_VEHICLES_N: r["distinct_vehicles"],
        COL_SLOTS_N:    r["slot_count"],
    } for r in rows])


def load_departures(conn, service_date: str, days: int) -> pd.DataFrame:
    cutoff = (date.fromisoformat(service_date) - timedelta(days=days)).isoformat()
    rows = conn.execute("""
        SELECT vd.service_date, vd.vehicle_no,
               r.line_code, r.descr AS route_descr, r.route_type,
               vd.departed_at, sa.slot_number, sa.scheduled_departure,
               sa.departure_deviation_mins
        FROM vehicle_departures vd
        LEFT JOIN routes r ON r.route_code = vd.route_code
        LEFT JOIN slot_assignments sa ON sa.trip_id = vd.trip_id
        WHERE vd.service_date >= ?
        ORDER BY vd.service_date DESC, vd.departed_at
    """, (cutoff,)).fetchall()
    return pd.DataFrame([{
        COL_DATE:       r["service_date"],
        COL_VEHICLE_NO: r["vehicle_no"],
        COL_LINE:       r["line_code"],
        COL_ROUTE:      r["route_descr"],
        COL_DIRECTION:  direction_gr(r["route_type"]),
        COL_DEPARTED:   r["departed_at"],
        COL_SLOT:       r["slot_number"],
        COL_SCHED_DEP:  r["scheduled_departure"],
        COL_DEV_MINS:   r["departure_deviation_mins"],
    } for r in rows])


def load_handoffs(conn, service_date: str, days: int) -> pd.DataFrame:
    cutoff = (date.fromisoformat(service_date) - timedelta(days=days)).isoformat()
    rows = conn.execute("""
        SELECT sh.service_date, r.line_code, r.descr AS route_descr,
               r.route_type, sh.slot_number,
               sh.outgoing_vehicle, sh.incoming_vehicle,
               sh.handoff_time, sh.gap_mins
        FROM slot_handoffs sh
        LEFT JOIN routes r ON r.route_code = sh.route_code
        WHERE sh.service_date >= ?
        ORDER BY sh.service_date DESC, sh.handoff_time
    """, (cutoff,)).fetchall()
    return pd.DataFrame([{
        COL_DATE:       r["service_date"],
        COL_LINE:       r["line_code"],
        COL_ROUTE:      r["route_descr"],
        COL_DIRECTION:  direction_gr(r["route_type"]),
        COL_SLOT:       r["slot_number"],
        COL_OUTGOING:   r["outgoing_vehicle"],
        COL_INCOMING:   r["incoming_vehicle"],
        COL_HANDOFF_T:  r["handoff_time"],
        COL_GAP:        r["gap_mins"],
    } for r in rows])


def load_reference(conn) -> pd.DataFrame:
    rows = conn.execute("""
        SELECT l.line_code, l.descr AS line_name,
               r.route_code, r.descr AS route_name, r.route_type,
               r.distance_m
        FROM lines l
        LEFT JOIN routes r ON r.line_code = l.line_code
        ORDER BY l.line_code
    """).fetchall()
    return pd.DataFrame([{
        COL_LINE:       r["line_code"],
        "Όνομα Γραμμής": r["line_name"],
        "Κωδικός Διαδρομής": r["route_code"],
        "Όνομα Διαδρομής":   r["route_name"],
        COL_DIRECTION:  direction_gr(r["route_type"]),
        COL_DISTANCE_M: r["distance_m"],
    } for r in rows])


# ── Formatting helpers ────────────────────────────────────────────────────────

def _style_header_row(ws, row: int = 1):
    for cell in ws[row]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER


def _autofit_columns(ws, max_width: int = 40):
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(
            max(length + 2, 8), max_width
        )


def _style_vehicles_sheet(ws):
    """Col A = blue, Col B = green, no header fill override."""
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font      = BODY_FONT
            cell.alignment = LEFT
        if len(row) >= 1:
            row[0].fill      = BLUE_FILL
            row[0].font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            row[0].alignment = CENTER
        if len(row) >= 2:
            row[1].fill = GREEN_FILL
            row[1].font = Font(bold=False, color="FFFFFF", name="Calibri", size=10)


def _style_grid_sheet(ws):
    """Time column left-aligned bold; vehicle cells centred."""
    GRID_HEADER = PatternFill("solid", start_color="1F3864")
    TIME_FILL   = PatternFill("solid", start_color="D9E1F2")
    VEH_FILL    = PatternFill("solid", start_color="E2EFDA")

    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        cell.fill      = GRID_HEADER
        cell.alignment = CENTER

    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            if i == 0:
                cell.font      = Font(bold=True, name="Calibri", size=9)
                cell.fill      = TIME_FILL
                cell.alignment = CENTER
            elif cell.value:
                cell.fill      = VEH_FILL
                cell.font      = Font(name="Calibri", size=9)
                cell.alignment = CENTER
            else:
                cell.font      = Font(name="Calibri", size=9)
                cell.alignment = CENTER


def _freeze_and_filter(ws, freeze: str = "B2"):
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


# ── Main export ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("output", nargs="?", default=None)
    parser.add_argument("--date", default=None, help="service date YYYY-MM-DD (default: latest in DB)")
    parser.add_argument("--days", type=int, default=14, help="days of detail history (default 14)")
    parser.add_argument("--all", action="store_true", help="full history for detail sheets")
    args = parser.parse_args()

    conn = db.get_connection()

    # Determine service date
    if args.date:
        service_date = args.date
    else:
        row = conn.execute(
            "SELECT MAX(service_date) d FROM daily_route_stats"
        ).fetchone()
        service_date = row["d"] if row and row["d"] else date.today().isoformat()

    detail_days = 9999 if args.all else args.days

    output_path = args.output or f"αθηνα_λεωφορεια_{service_date}.xlsx"

    print(f"Εξαγωγή για ημερομηνία: {service_date}")
    print(f"Ιστορικό λεπτομερειών: {'όλο' if args.all else f'{detail_days} ημέρες'}")

    print("Φόρτωση δεδομένων...")
    df_vehicles  = load_vehicles_per_day(conn, service_date)
    df_grid      = load_departure_grid(conn, service_date)
    df_trips     = load_trips_with_stops(conn, service_date, detail_days)
    df_stats     = load_daily_stats(conn)
    df_departures = load_departures(conn, service_date, detail_days)
    df_handoffs  = load_handoffs(conn, service_date, detail_days)
    df_reference = load_reference(conn)
    conn.close()

    print(f"Εγγραφή {output_path} ...")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet 1: Vehicles per day (blue/green columns)
        df_vehicles.to_excel(writer, sheet_name=SHEET_NAMES["vehicles"], index=False)

        # Sheet 2: Departure grid
        df_grid.to_excel(writer, sheet_name=SHEET_NAMES["grid"])

        # Sheet 3: Trips with stop times
        df_trips.to_excel(writer, sheet_name=SHEET_NAMES["trips"], index=False)

        # Sheet 4: Daily stats
        df_stats.to_excel(writer, sheet_name=SHEET_NAMES["stats"], index=False)

        # Sheet 5: Departures
        df_departures.to_excel(writer, sheet_name=SHEET_NAMES["departures"], index=False)

        # Sheet 6: Handoffs
        df_handoffs.to_excel(writer, sheet_name=SHEET_NAMES["handoffs"], index=False)

        # Sheet 7: Reference
        df_reference.to_excel(writer, sheet_name=SHEET_NAMES["reference"], index=False)

    # Post-process formatting
    wb = load_workbook(output_path)

    # Sheet 1 styling
    ws1 = wb[SHEET_NAMES["vehicles"]]
    _style_header_row(ws1)
    _style_vehicles_sheet(ws1)
    ws1.freeze_panes = "A2"
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 40

    # Sheet 2 styling (grid)
    ws2 = wb[SHEET_NAMES["grid"]]
    _style_grid_sheet(ws2)
    ws2.freeze_panes = "B2"
    # Set narrow column widths for route columns
    for col in ws2.iter_cols(min_col=2):
        letter = col[0].column_letter
        ws2.column_dimensions[letter].width = 14
    ws2.column_dimensions["A"].width = 8

    # Remaining sheets
    for key in ("trips", "stats", "departures", "handoffs", "reference"):
        ws = wb[SHEET_NAMES[key]]
        _style_header_row(ws)
        _freeze_and_filter(ws)
        _autofit_columns(ws)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font      = BODY_FONT
                cell.alignment = LEFT

    wb.save(output_path)

    print(f"\n✓ Αρχείο αποθηκεύτηκε: {output_path}")
    print(f"  {SHEET_NAMES['vehicles']}: {len(df_vehicles)} γραμμές")
    print(f"  {SHEET_NAMES['grid']}: {len(df_grid)} χρονικές γραμμές × {len(df_grid.columns)} διαδρομές")
    print(f"  {SHEET_NAMES['trips']}: {len(df_trips)} γραμμές")
    print(f"  {SHEET_NAMES['stats']}: {len(df_stats)} γραμμές")
    print(f"  {SHEET_NAMES['departures']}: {len(df_departures)} γραμμές")
    print(f"  {SHEET_NAMES['handoffs']}: {len(df_handoffs)} γραμμές")
    print(f"  {SHEET_NAMES['reference']}: {len(df_reference)} γραμμές")


if __name__ == "__main__":
    main()
